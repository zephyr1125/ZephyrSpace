#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
生成护城河评估结果的Excel汇总表
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "护城河评估结果"

# 设置列宽
col_widths = {
    'A': 18,
    'B': 12,
    'C': 12,
    'D': 14,
    'E': 18,
    'F': 16,
    'G': 20,
    'H': 14,
    'I': 20
}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# 定义颜色
header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
strong_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿色
weak_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # 黄色
danger_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # 红色
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 标题
ws.merge_cells('A1:I1')
title = ws['A1']
title.value = "第四轮筛选评估 - 护城河强度分析结果表"
title.font = Font(bold=True, size=14, color="FFFFFF")
title.fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
title.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 25

# 评估日期
ws['A2'].value = "评估日期：2026年4月30日"
ws['A2'].font = Font(italic=True, size=9)

# 表头
headers = [
    "公司名称",
    "代码",
    "市值(亿元)",
    "毛利率(%)",
    "竞争优势",
    "护城河等级",
    "主要负面事件",
    "决策",
    "优先级"
]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# 数据行
data = [
    {
        "name": "TCL智家",
        "code": "002668",
        "market_cap": "103.5",
        "margin": "11.59",
        "advantages": "品牌✓/技术✓\n成本✓/渠道✓",
        "moat": "⭐弱",
        "risks": "高榕诉讼2352万+\n内控风险",
        "decision": "🟡待评估",
        "priority": "低"
    },
    {
        "name": "爱尔眼科",
        "code": "300015",
        "market_cap": "1000",
        "margin": "48.12",
        "advantages": "品牌✓✓✓\n技术✓✓✓\n网络✓✓✓✓",
        "moat": "⭐⭐⭐强",
        "risks": ">400例医保处罚\n结构性问题",
        "decision": "✅保留",
        "priority": "高"
    },
    {
        "name": "艾力斯",
        "code": "688578",
        "market_cap": "440",
        "margin": "95.97",
        "advantages": "技术✓✓✓✓\n研发✓✓✓\n产品壁垒✓✓✓",
        "moat": "⭐⭐⭐强",
        "risks": "复星仲裁2.55亿\n竞争压力",
        "decision": "✅保留",
        "priority": "高"
    },
    {
        "name": "紫金矿业",
        "code": "601899",
        "market_cap": "8854.71",
        "margin": "36.33",
        "advantages": "成本✓✓✓\n规模✓✓✓✓\n资源✓✓✓",
        "moat": "⭐⭐⭐强",
        "risks": "ICSID仲裁4.3亿$\n海外风险",
        "decision": "✅保留",
        "priority": "高"
    }
]

# 填充数据
for row_num, item in enumerate(data, 5):
    col = 1
    values = [
        item["name"],
        item["code"],
        item["market_cap"],
        item["margin"],
        item["advantages"],
        item["moat"],
        item["risks"],
        item["decision"],
        item["priority"]
    ]
    
    for value in values:
        cell = ws.cell(row=row_num, column=col)
        cell.value = value
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = border
        
        # 根据内容设置背景色
        if col == 6:  # 护城河等级
            if "强" in str(value):
                cell.fill = strong_fill
            elif "弱" in str(value):
                cell.fill = weak_fill
        elif col == 8:  # 决策
            if "待评估" in str(value):
                cell.fill = weak_fill
        
        col += 1
    
    ws.row_dimensions[row_num].height = 40

# 添加说明
summary_row = 9
ws.merge_cells(f'A{summary_row}:I{summary_row}')
summary = ws[f'A{summary_row}']
summary.value = "【第四轮筛选结论】"
summary.font = Font(bold=True, size=11)
summary.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

detail_text = [
    "✅ 必须保留（3家）：爱尔眼科、艾力斯、紫金矿业 - 符合护城河⭐⭐⭐强或毛利率上升标准",
    "🟡 需要评估（1家）：TCL智家 - 毛利率11.59%低于标准，高榕诉讼侵蚀利润，建议明年H1评估",
    "📌 关键指标：市值≥50亿✓  毛利率>25%（或上升趋势）✓  竞争优势≥2个✓"
]

for idx, text in enumerate(detail_text, summary_row + 1):
    cell = ws[f'A{idx}']
    cell.value = text
    cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(f'A{idx}:I{idx}')
    ws.row_dimensions[idx].height = 20

# 保存文件
output_file = os.path.join(os.path.dirname(__file__), '..', '4轮筛选护城河评估汇总表.xlsx')
wb.save(output_file)
print(f"✓ Excel汇总表已生成：{output_file}")
